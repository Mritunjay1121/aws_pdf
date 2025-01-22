from flask import Flask, render_template, request, jsonify, send_file
import boto3
import pandas as pd
from datetime import datetime
from config import AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, AWS_REGION, BUCKET_NAME
from botocore.exceptions import ClientError
from io import BytesIO
from trp import Document
from PyPDF2 import PdfReader, PdfWriter
import time
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
import re
app = Flask(__name__)

# Initialize AWS Textract client
client = boto3.client('textract',
                     aws_access_key_id=AWS_ACCESS_KEY_ID,
                     aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                     region_name=AWS_REGION)

def upload_to_s3(file_bytes, original_filename):
    s3 = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                      aws_secret_access_key=AWS_SECRET_ACCESS_KEY, region_name=AWS_REGION)
    bucket_name = BUCKET_NAME
    date_str = datetime.now().strftime("%Y-%m-%d")
    base_name, extension = original_filename.rsplit('.', 1)
    sanitized_base_name = base_name.replace(' ', '_')
    file_name = f'{sanitized_base_name}_{date_str}.{extension}'
    try:
        s3.upload_fileobj(BytesIO(file_bytes), bucket_name, file_name)
        return bucket_name, file_name
    except ClientError as e:
        print(f"Error uploading file to S3: {e}")
        return None, None

def extract_text_from_pdf(file_bytes, original_filename):
    bucket_name, file_name = upload_to_s3(file_bytes, original_filename)
    if not bucket_name or not file_name:
        return None
    
    try:
        response = client.start_document_analysis(
            DocumentLocation={
                'S3Object': {
                    'Bucket': bucket_name,
                    'Name': file_name
                }
            },
            FeatureTypes=['TABLES', 'FORMS']
        )
        job_id = response['JobId']
        
        while True:
            response = client.get_document_analysis(JobId=job_id)
            status = response['JobStatus']
            if status in ['SUCCEEDED', 'FAILED']:
                break
            time.sleep(5)
        
        if status == 'FAILED':
            raise Exception("Document analysis failed")
        
        # Collect all pages of the response
        all_blocks = response['Blocks']
        next_token = response.get('NextToken', None)
        
        while next_token:
            response = client.get_document_analysis(JobId=job_id, NextToken=next_token)
            all_blocks.extend(response['Blocks'])
            next_token = response.get('NextToken', None)
        
        response['Blocks'] = all_blocks
        return response
    
    except ClientError as e:
        print(f"AWS Textract error: {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None

def extract_text_from_image(file_bytes):
    try:
        response = client.detect_document_text(Document={'Bytes': file_bytes})
        lines = [item['Text'] for item in response['Blocks'] if item['BlockType'] == 'LINE']
        return lines
    
    except ClientError as e:
        print(f"AWS Textract error: {e}")
        return []
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return []

def extract_text_from_image_table(file_bytes):
    try:
        response = client.detect_document_text(Document={'Bytes': file_bytes})
        blocks = response['Blocks']
        if not blocks:
            return []
        
        # Group blocks by their detected table IDs
        table_blocks = {}
        for block in blocks:
            if block['BlockType'] == 'TABLE':
                if 'Table' not in table_blocks:
                    table_blocks['Table'] = []
                table_blocks['Table'].append(block)
        
        if 'Table' not in table_blocks:
            return []
        
        # Process each detected table and extract text
        tables = []
        for table_block in table_blocks['Table']:
            rows = {}
            for relationship in table_block['Relationships']:
                if relationship['Type'] == 'CHILD':
                    for child_id in relationship['Ids']:
                        cell = blocks[child_id]
                        if cell['BlockType'] == 'CELL':
                            row_index = cell['RowIndex']
                            column_index = cell['ColumnIndex']
                            if row_index not in rows:
                                rows[row_index] = {}
                            rows[row_index][column_index] = cell['Text']
            
            # Convert rows into DataFrame for each table
            table_data = []
            for row_index in sorted(rows.keys()):
                row_data = [rows[row_index].get(column_index, '') for column_index in sorted(rows[row_index].keys())]
                table_data.append(row_data)
            
            df = pd.DataFrame(table_data)
            tables.append(df)
        
        return tables
    
    except ClientError as e:
        print(f"AWS Textract error: {e}")
        return []
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return []


def clean_and_type_check(value):
    
    if pd.isna(value) or not isinstance(value, str):
        value = str(value) if value is not None else ''
    cleaned_value = re.sub(r'[,\[\]\(\)\$]', '', value)

    
    
    is_number = False
    
    try:
        
        float(cleaned_value)
        is_number = True
    except ValueError:
       
        pass  

    return cleaned_value, is_number


def process_textract_response(response):
    try:
        if 'Blocks' not in response:
            raise ValueError("Response does not contain 'Blocks'")

        doc = Document(response)
        tables = []
        
        for page in doc.pages:
            for table in page.tables:
                table_data = []
                for row in table.rows:
                    row_data = [cell.text if cell.text else "" for cell in row.cells]
                    table_data.append(row_data)
                if table_data:  # Check if table_data has content
                    df = pd.DataFrame(table_data)
                    df.fillna(0,inplace=True)

                    df = df.replace(r'^\s*-\s*$', 0, regex=True)
                    all_int=True
                    for col in df.columns:
                        for i in range(len(df[col])):
                        
                            val, is_number = clean_and_type_check(df.loc[i, col])
                            if is_number==True:
                                df.loc[i, col] = float(val)
                            elif is_number==False:
                                all_int=False
                        if all_int==True:
                            df[col] = pd.to_numeric(df[col])

                    # df.to_excel("a_tablenew.xlsx",index=False)



                    tables.append({
                        'columns': df.columns.tolist(),
                        'data': df.to_dict('records')
                    })

        return tables

    except Exception as e:
        print(f"Error processing Textract response: {e}")
        return []  # Return an empty list if there's an error

def merge_pages_from_multiple_pdfs(files_and_pages):
    """
    files_and_pages: list of tuples [(file_bytes, [page_numbers]), ...]
    """
    try:
        pdf_writer = PdfWriter()
        
        for file_bytes, page_numbers in files_and_pages:
            pdf_reader = PdfReader(BytesIO(file_bytes))
            total_pages = len(pdf_reader.pages)
            
            # for page_num in page_numbers:
            #     if page_num < 0 or page_num >= total_pages:
            #         raise ValueError(f"Invalid page number {page_num + 1}. File has {total_pages} pages.")
            
            # for page_num in page_numbers:
            #     pdf_writer.add_page(pdf_reader.pages[page_num])

            if page_numbers == [-1]:  # If -1 is given for all pages
                page_numbers = list(range(total_pages))
            else:
                for page_num in page_numbers:
                    if page_num < -1 or page_num >= total_pages:
                        raise ValueError(f"Invalid page number {page_num + 1}. File has {total_pages} pages.")
            
            for page_num in page_numbers:
                if page_num != -1:  # Skip if we're already using all pages
                    pdf_writer.add_page(pdf_reader.pages[page_num])
                else:  # Add all pages
                    for page in pdf_reader.pages:
                        pdf_writer.add_page(page)

        
        output_bytes = BytesIO()
        pdf_writer.write(output_bytes)
        output_bytes.seek(0)
        
        return output_bytes
        
    except Exception as e:
        raise Exception(f"Error in PDF processing: {str(e)}")

@app.route('/')
def index():
    return render_template('index.html')
@app.route('/upload', methods=['POST'])
def upload():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    files = request.files.getlist('files[]')
    if not files or files[0].filename == '':
        return jsonify({'error': 'No files selected'}), 400

    try:
        pages_str = request.form.get('pages', '')
        files_and_pages = []
        
        if pages_str:
            pages_per_file = pages_str.split('|')
            if len(pages_per_file) != len(files):
                return jsonify({'error': 'Number of page sets does not match number of files'}), 400
            
            for file, pages in zip(files, pages_per_file):
                if not file.filename.endswith('.pdf'):
                    return jsonify({'error': f'File {file.filename} must be a PDF'}), 400
                
                try:
                    if pages == '-1':
                        page_numbers = [-1]  # All pages
                    else:
                        page_numbers = [int(p.strip()) - 1 for p in pages.split(',')]

                    # page_numbers = [int(p.strip()) - 1 for p in pages.split(',')]
                    files_and_pages.append((file.read(), page_numbers))
                except ValueError:
                    return jsonify({'error': 'Invalid page numbers format'}), 400
        else:
            for file in files:
                if not file.filename.endswith('.pdf'):
                    return jsonify({'error': f'File {file.filename} must be a PDF'}), 400
                
                file_bytes = file.read()
                pdf_reader = PdfReader(BytesIO(file_bytes))

                page_numbers = [-1]
                # page_numbers = list(range(len(pdf_reader.pages)))
                files_and_pages.append((file_bytes, page_numbers))

        merged_pdf_bytes = merge_pages_from_multiple_pdfs(files_and_pages)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_filename = f"merged_{timestamp}.pdf"

        response = extract_text_from_pdf(merged_pdf_bytes.getvalue(), merged_filename)
        if response:
            tables = process_textract_response(response)
            if tables:
                return jsonify({'tables': tables})
            else:
                return jsonify({'error': 'No tables were detected in the document.'}), 400
        else:
            return jsonify({'error': 'Failed to process document.'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 400




@app.route('/download_all_tables', methods=['POST'])
def download_all_tables():
    try:
        data = request.json
        all_tables = data['tables']
        format_type = data.get('format', 'one')  # Default to one sheet if not specified
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")


        def adjust_column_widths(worksheet, df):
            for col in df.columns:
                column = get_column_letter(df.columns.get_loc(col) + 1)
                max_length = 0
                column_widths = []
                
                for cell in worksheet[column]:
                    if cell.value:
                        length = len(str(cell.value))
                        column_widths.append(length)
                        if length > max_length:
                            max_length = length
                
                # The width is set to a little larger than the longest cell content for better readability
                adjusted_width = max(max_length, 10)  # Minimum width of 10 for better visibility
                worksheet.column_dimensions[column].width = adjusted_width





        
        def format_numbers(df, worksheet):
            for col in df.select_dtypes(include=['float64', 'int64']):
                col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                for row in range(2, worksheet.max_row + 1):  # Start from row 2 as row 1 is header
                    cell = worksheet[f'{col_letter}{row}']
                    
                    # Apply Number format
                    cell.number_format = numbers.FORMAT_NUMBER_00  # For Number format with two decimal places
                    
                    # Apply Accounting format
                    # cell.number_format = numbers.FORMAT_ACCOUNTING_USD  # For US Dollar accounting format
                    if cell.value is not None and cell.value != '':
                        cell.value = float(cell.value)  # Ensure the value is a float for accounting format

        if format_type == 'one':
            with pd.ExcelWriter(f'all_tables_{timestamp}.xlsx', engine='openpyxl') as writer:
                for i, table in enumerate(all_tables):
                    df = pd.DataFrame(table['data'])
                    
                    # table_name = f'Table_{i+1}'
                    table_name=table.get('name')

                    sheet_name = table_name[:]
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                    
                    # Remove gridlines and apply number format
                    workbook = writer.book
                    worksheet = workbook[sheet_name]
                    worksheet.sheet_view.showGridLines = False
                    format_numbers(df, worksheet)
                    adjust_column_widths(worksheet, df)

        else:  # Multiple files for each table
            for i, table in enumerate(all_tables):
                df = pd.DataFrame(table['data'])
                
                table_name = table.get('name', f'Table_{i+1}')
                filename = f"{table_name}.xlsx"

                # filename = f'table_{i+1}_{timestamp}.xlsx'
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
                    
                    # Remove gridlines and apply number format
                    workbook = writer.book
                    worksheet = workbook['Sheet1']
                    worksheet.sheet_view.showGridLines = False
                    format_numbers(df, worksheet)
                    adjust_column_widths(worksheet, df)

        return jsonify({
            'success': True,
            'message': f'Successfully saved {len(all_tables)} tables'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 400




@app.route('/modify_table', methods=['POST'])
def modify_table():
    try:
        data = request.json
        table_index = data['table_index']
        action = data['action']
        table_data = data['table_data']
        
        df = pd.DataFrame(table_data['data'])
        
        if action == 'delete_column':
            column_name = data['column_name']
            if len(df.columns) <= 1:
                return jsonify({'error': 'Cannot delete the last column'}), 400
            df = df.drop(columns=[column_name])
            
        # elif action == 'add_column':
        #     column_name = data['column_name']
        #     column_index = data.get('column_index', len(df.columns))
        #     if column_name in df.columns:
        #         return jsonify({'error': 'Column already exists'}), 400
        #     df.insert(column_index, column_name, '')
            



        elif action == 'add_column':
            column_name = data['column_name']
            column_index = data.get('column_index', len(df.columns))
            if column_name in df.columns:
                return jsonify({'error': 'Column already exists'}), 400
            df.insert(column_index, column_name, '')
            updated_table = {
                'columns': df.columns.tolist(),
                'data': df.to_dict('records')
            }
            
        
            


       
            

        elif action == 'delete_row':
            row_index = data['row_index']
            if len(df) <= 1:
                return jsonify({'error': 'Cannot delete the last row'}), 400
            df = df.drop(index=row_index).reset_index(drop=True)
            
        elif action == 'add_row':
            row_index = data.get('row_index', len(df))
            new_row = pd.DataFrame([['' for _ in df.columns]], columns=df.columns)
            if row_index == len(df):
                df = pd.concat([df, new_row], ignore_index=True)
            else:
                df = pd.concat([df.iloc[:row_index], new_row, df.iloc[row_index:]]).reset_index(drop=True)

        # elif action == 'edit_cell':
        #     row_index = data['row_index']
        #     column_name = data['column_name']
        #     new_value = data['new_value']
        #     df.at[row_index, column_name] = new_value
                
        elif action == 'edit_cell':
            row_index = data['row_index']
            column_name = data['column_name']
            new_value = data['new_value']
            
            # Determine if the column is numeric
            if column_name in df.columns:
                if pd.api.types.is_numeric_dtype(df[column_name]):
                    try:
                        # Try to convert new_value to float
                        df.at[row_index, column_name] = float(new_value)
                    except ValueError:
                        # If conversion fails, keep as string
                        df.at[row_index, column_name] = new_value
                else:
                    # If the column is not numeric (i.e., object), set as string
                    df.at[row_index, column_name] = str(new_value)

        elif action == 'edit_column_name':
            old_column_name = data['old_column_name']
            new_column_name = data['new_column_name']
            if new_column_name in df.columns:
                return jsonify({'error': 'A column with this name already exists'}), 400
            df = df.rename(columns={old_column_name: new_column_name})
        elif action == 'adjust_numeric':
            column_name = data['column_name']
            operation = data['operation']
            factor = data['factor']
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(lambda x: float(x) if pd.notna(x) else x)
                if operation == 'multiply':
                    df[column_name] = df[column_name] * factor
                elif operation == 'divide':
                    df[column_name] = df[column_name] / factor
        else:
            return jsonify({'error': 'Unknown action'}), 400

        updated_table = {
            'columns': df.columns.tolist(),
            'data': df.to_dict('records')
        }
        
        return jsonify({
            'success': True,
            'table': updated_table
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# @app.route('/save_table', methods=['POST'])
# def save_table():
#     try:
#         data = request.json
#         table_index = data['table_index']
#         table_index+=1
#         table_data = data['table_data']
        
#         df = pd.DataFrame(table_data['data'])
        
#         filename = f'table_{table_index}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
#         df.to_excel(filename, index=False)
        
#         return jsonify({'success': True, 'filename': filename})
#     except Exception as e:
#         return jsonify({'error': str(e)}), 400

@app.route('/save_table', methods=['POST'])
def save_table():
    try:
        data = request.json
        table_data = data['table_data']
        
        # Extract the DataFrame from the table data
        df = pd.DataFrame(table_data['data'])
        
        # Get the table name from the table data, defaulting to a timestamped name if not available
        table_name = table_data.get('name')
        
        # Sanitize the table name to make it safe for filenames
        sanitized_table_name = ''.join(c if c.isalnum() or c in (' ', '.', '_') else '_' for c in table_name)
        
        # Ensure the filename is not too long (maximum filename length varies by OS, commonly 255 characters)
        max_filename_length = 255
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'{sanitized_table_name}.xlsx'
        # if len(filename) > max_filename_length:
        #     # Truncate the sanitized_table_name to fit within the limit
        #     allowed_name_length = max_filename_length - len(f'_{timestamp}.xlsx')
        #     sanitized_table_name = sanitized_table_name[:allowed_name_length]
        #     filename = f'{sanitized_table_name}.xlsx'
        
        # Save the DataFrame to an Excel file with the sanitized filename
        df.to_excel(filename, index=False)
        
        return jsonify({'success': True, 'filename': filename})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

if __name__ == '__main__':
    app.run(debug=True)